import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def build_netflix_excel():
    excel_path = os.path.join('Excel', 'Netflix_Analytics.xlsx')
    raw_csv = os.path.join('Dataset', 'netflix_titles_raw.csv')
    cleaned_csv = os.path.join('Dataset', 'netflix_titles_cleaned.csv')

    print("Loading datasets for Excel...")
    df_raw = pd.read_csv(raw_csv)
    df_clean = pd.read_csv(cleaned_csv)

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=9, italic=True, color="CCCCCC")
    section_font = Font(name=font_family, size=11, bold=True, color="1F2937")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    kpi_val_font = Font(name=font_family, size=18, bold=True, color="E50914")
    kpi_lbl_font = Font(name=font_family, size=8, bold=True, color="4B5563")
    data_font = Font(name=font_family, size=9, color="1F2937")
    total_font = Font(name=font_family, size=9, bold=True, color="111827")

    dark_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    header_fill = PatternFill(start_color="221F1F", end_color="221F1F", fill_type="solid")
    sub_header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    kpi_bg_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    total_row_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=Side(border_style="double", color="111827"))

    # ----------------------------------------------------
    # 1. SHEET: KPI_Summary
    # ----------------------------------------------------
    print("Building KPI_Summary sheet...")
    ws_kpi = wb.create_sheet(title="KPI_Summary")
    ws_kpi.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_kpi.merge_cells("B2:K2")
    ws_kpi["B2"] = "NETFLIX CONTENT ANALYTICS - EXECUTIVE SUMMARY"
    ws_kpi["B2"].font = title_font
    ws_kpi["B2"].fill = dark_fill
    ws_kpi["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[2].height = 28

    ws_kpi.merge_cells("B3:K3")
    ws_kpi["B3"] = "Catalog composition, multi-year content additions, runtime benchmarks, and maturity ratings"
    ws_kpi["B3"].font = subtitle_font
    ws_kpi["B3"].fill = dark_fill
    ws_kpi["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[3].height = 16

    # KPI Cards Layout (Row 5 - 6)
    kpis = [
        ("TOTAL TITLES", "=COUNTA(Cleaned_Data!A2:A8808)", "#,##0", "B", "C"),
        ("TOTAL MOVIES", '=COUNTIF(Cleaned_Data!B2:B8808, "Movie")', "#,##0", "D", "E"),
        ("TOTAL TV SHOWS", '=COUNTIF(Cleaned_Data!B2:B8808, "TV Show")', "#,##0", "F", "G"),
        ("MOVIE SHARE", "=D6/B6", "0.0%", "H", "H"),
        ("TV SHOW SHARE", "=F6/B6", "0.0%", "I", "I"),
        ("AVG MOVIE RUNTIME", "=AVERAGE(Cleaned_Data!P2:P8808)", '0.0" min"', "J", "K")
    ]

    for lbl, formula, num_fmt, start_col, end_col in kpis:
        cell_lbl = f"{start_col}5"
        if start_col != end_col:
            ws_kpi.merge_cells(f"{start_col}5:{end_col}5")
        ws_kpi[cell_lbl] = lbl
        ws_kpi[cell_lbl].font = kpi_lbl_font
        ws_kpi[cell_lbl].fill = kpi_bg_fill
        ws_kpi[cell_lbl].alignment = Alignment(horizontal="center", vertical="center")

        cell_val = f"{start_col}6"
        if start_col != end_col:
            ws_kpi.merge_cells(f"{start_col}6:{end_col}6")
        ws_kpi[cell_val] = formula
        ws_kpi[cell_val].font = kpi_val_font
        ws_kpi[cell_val].fill = kpi_bg_fill
        ws_kpi[cell_val].alignment = Alignment(horizontal="center", vertical="center")
        ws_kpi[cell_val].number_format = num_fmt

        for col_letter in [start_col, end_col]:
            for r in [5, 6]:
                ws_kpi[f"{col_letter}{r}"].border = thin_border

    ws_kpi.row_dimensions[5].height = 16
    ws_kpi.row_dimensions[6].height = 28

    # Section 1: Annual Additions Breakdown
    ws_kpi["B8"] = "1. Annual Content Ingestion Trend (2012 - 2021)"
    ws_kpi["B8"].font = section_font

    annual_headers = ["Year Added", "Movies Added", "TV Shows Added", "Total Titles Added", "Movie % Share", "YoY Growth"]
    for col_idx, h in enumerate(annual_headers, start=2):
        col_let = get_column_letter(col_idx)
        cell = f"{col_let}9"
        ws_kpi[cell] = h
        ws_kpi[cell].font = header_font
        ws_kpi[cell].fill = header_fill
        ws_kpi[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_kpi[cell].border = thin_border
    ws_kpi.row_dimensions[9].height = 20

    years = [int(y) for y in sorted(df_clean['year_added'].dropna().unique()) if y >= 2012]
    curr_row = 10
    for y_int in years:
        ws_kpi[f"B{curr_row}"] = y_int
        ws_kpi[f"B{curr_row}"].font = data_font
        ws_kpi[f"B{curr_row}"].alignment = Alignment(horizontal="center")
        ws_kpi[f"B{curr_row}"].border = thin_border

        ws_kpi[f"C{curr_row}"] = f'=COUNTIFS(Cleaned_Data!$I$2:$I$8808, {y_int}, Cleaned_Data!$B$2:$B$8808, "Movie")'
        ws_kpi[f"C{curr_row}"].font = data_font
        ws_kpi[f"C{curr_row}"].number_format = "#,##0"
        ws_kpi[f"C{curr_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"C{curr_row}"].border = thin_border

        ws_kpi[f"D{curr_row}"] = f'=COUNTIFS(Cleaned_Data!$I$2:$I$8808, {y_int}, Cleaned_Data!$B$2:$B$8808, "TV Show")'
        ws_kpi[f"D{curr_row}"].font = data_font
        ws_kpi[f"D{curr_row}"].number_format = "#,##0"
        ws_kpi[f"D{curr_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"D{curr_row}"].border = thin_border

        ws_kpi[f"E{curr_row}"] = f"=C{curr_row}+D{curr_row}"
        ws_kpi[f"E{curr_row}"].font = data_font
        ws_kpi[f"E{curr_row}"].number_format = "#,##0"
        ws_kpi[f"E{curr_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"E{curr_row}"].border = thin_border

        ws_kpi[f"F{curr_row}"] = f"=C{curr_row}/E{curr_row}"
        ws_kpi[f"F{curr_row}"].font = data_font
        ws_kpi[f"F{curr_row}"].number_format = "0.0%"
        ws_kpi[f"F{curr_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"F{curr_row}"].border = thin_border

        if curr_row == 10:
            ws_kpi[f"G{curr_row}"] = "-"
        else:
            ws_kpi[f"G{curr_row}"] = f"=(E{curr_row}-E{curr_row-1})/E{curr_row-1}"
            ws_kpi[f"G{curr_row}"].number_format = "0.0%"
        ws_kpi[f"G{curr_row}"].font = data_font
        ws_kpi[f"G{curr_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"G{curr_row}"].border = thin_border

        if curr_row % 2 == 1:
            for c in ["B", "C", "D", "E", "F", "G"]:
                ws_kpi[f"{c}{curr_row}"].fill = alt_row_fill
        curr_row += 1

    # Total Row for Annual
    ws_kpi[f"B{curr_row}"] = "Total (2012-2021)"
    ws_kpi[f"B{curr_row}"].font = total_font
    ws_kpi[f"B{curr_row}"].fill = total_row_fill
    ws_kpi[f"B{curr_row}"].border = double_bottom_border

    ws_kpi[f"C{curr_row}"] = f"=SUM(C10:C{curr_row-1})"
    ws_kpi[f"C{curr_row}"].font = total_font
    ws_kpi[f"C{curr_row}"].number_format = "#,##0"
    ws_kpi[f"C{curr_row}"].fill = total_row_fill
    ws_kpi[f"C{curr_row}"].border = double_bottom_border

    ws_kpi[f"D{curr_row}"] = f"=SUM(D10:D{curr_row-1})"
    ws_kpi[f"D{curr_row}"].font = total_font
    ws_kpi[f"D{curr_row}"].number_format = "#,##0"
    ws_kpi[f"D{curr_row}"].fill = total_row_fill
    ws_kpi[f"D{curr_row}"].border = double_bottom_border

    ws_kpi[f"E{curr_row}"] = f"=SUM(E10:E{curr_row-1})"
    ws_kpi[f"E{curr_row}"].font = total_font
    ws_kpi[f"E{curr_row}"].number_format = "#,##0"
    ws_kpi[f"E{curr_row}"].fill = total_row_fill
    ws_kpi[f"E{curr_row}"].border = double_bottom_border

    ws_kpi[f"F{curr_row}"] = f"=C{curr_row}/E{curr_row}"
    ws_kpi[f"F{curr_row}"].font = total_font
    ws_kpi[f"F{curr_row}"].number_format = "0.0%"
    ws_kpi[f"F{curr_row}"].fill = total_row_fill
    ws_kpi[f"F{curr_row}"].border = double_bottom_border

    ws_kpi[f"G{curr_row}"] = "-"
    ws_kpi[f"G{curr_row}"].font = total_font
    ws_kpi[f"G{curr_row}"].alignment = Alignment(horizontal="center")
    ws_kpi[f"G{curr_row}"].fill = total_row_fill
    ws_kpi[f"G{curr_row}"].border = double_bottom_border

    # Section 2: Ratings & Maturity Audience Profile
    ws_kpi["I8"] = "2. Target Audience Maturity Profile"
    ws_kpi["I8"].font = section_font

    rating_headers = ["Audience Bucket", "Ratings Included", "Titles Count", "% of Catalog"]
    for idx, rh in enumerate(rating_headers):
        col_let = chr(ord('I') + idx)
        cell = f"{col_let}9"
        ws_kpi[cell] = rh
        ws_kpi[cell].font = header_font
        ws_kpi[cell].fill = header_fill
        ws_kpi[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_kpi[cell].border = thin_border

    buckets = [
        ("Adults (18+)", "TV-MA, R, NC-17", '=COUNTIF(Cleaned_Data!N2:N8808, "Adults (18+)")'),
        ("Teens (13-17)", "TV-14, PG-13", '=COUNTIF(Cleaned_Data!N2:N8808, "Teens (13-17)")'),
        ("Older Kids (7-12)", "TV-PG, PG, TV-Y7, TV-Y7-FV", '=COUNTIF(Cleaned_Data!N2:N8808, "Older Kids (7-12)")'),
        ("Kids (All Ages)", "TV-Y, TV-G, G", '=COUNTIF(Cleaned_Data!N2:N8808, "Kids (All Ages)")'),
        ("Unrated / NR", "NR", '=COUNTIF(Cleaned_Data!N2:N8808, "Unrated / NR")')
    ]

    r_row = 10
    for b_name, b_ratings, b_form in buckets:
        ws_kpi[f"I{r_row}"] = b_name
        ws_kpi[f"I{r_row}"].font = data_font
        ws_kpi[f"I{r_row}"].border = thin_border

        ws_kpi[f"J{r_row}"] = b_ratings
        ws_kpi[f"J{r_row}"].font = data_font
        ws_kpi[f"J{r_row}"].border = thin_border

        ws_kpi[f"K{r_row}"] = b_form
        ws_kpi[f"K{r_row}"].font = data_font
        ws_kpi[f"K{r_row}"].number_format = "#,##0"
        ws_kpi[f"K{r_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"K{r_row}"].border = thin_border

        ws_kpi[f"L{r_row}"] = f"=K{r_row}/B6"
        ws_kpi[f"L{r_row}"].font = data_font
        ws_kpi[f"L{r_row}"].number_format = "0.0%"
        ws_kpi[f"L{r_row}"].alignment = Alignment(horizontal="right")
        ws_kpi[f"L{r_row}"].border = thin_border

        if r_row % 2 == 1:
            for c in ["I", "J", "K", "L"]:
                ws_kpi[f"{c}{r_row}"].fill = alt_row_fill
        r_row += 1

    # Total Row for Maturity
    ws_kpi[f"I{r_row}"] = "Total Catalog"
    ws_kpi[f"I{r_row}"].font = total_font
    ws_kpi[f"I{r_row}"].fill = total_row_fill
    ws_kpi[f"I{r_row}"].border = double_bottom_border

    ws_kpi[f"J{r_row}"] = "All Certifications"
    ws_kpi[f"J{r_row}"].font = total_font
    ws_kpi[f"J{r_row}"].fill = total_row_fill
    ws_kpi[f"J{r_row}"].border = double_bottom_border

    ws_kpi[f"K{r_row}"] = f"=SUM(K10:K{r_row-1})"
    ws_kpi[f"K{r_row}"].font = total_font
    ws_kpi[f"K{r_row}"].number_format = "#,##0"
    ws_kpi[f"K{r_row}"].alignment = Alignment(horizontal="right")
    ws_kpi[f"K{r_row}"].fill = total_row_fill
    ws_kpi[f"K{r_row}"].border = double_bottom_border

    ws_kpi[f"L{r_row}"] = f"=SUM(L10:L{r_row-1})"
    ws_kpi[f"L{r_row}"].font = total_font
    ws_kpi[f"L{r_row}"].number_format = "0.0%"
    ws_kpi[f"L{r_row}"].alignment = Alignment(horizontal="right")
    ws_kpi[f"L{r_row}"].fill = total_row_fill
    ws_kpi[f"L{r_row}"].border = double_bottom_border

    kpi_widths = {"A": 4, "B": 18, "C": 16, "D": 16, "E": 18, "F": 16, "G": 16, "H": 16, "I": 20, "J": 26, "K": 16, "L": 16}
    for col, width in kpi_widths.items():
        ws_kpi.column_dimensions[col].width = width

    # ----------------------------------------------------
    # 2. SHEET: Pivot_Analysis
    # ----------------------------------------------------
    print("Building Pivot_Analysis sheet...")
    ws_pivot = wb.create_sheet(title="Pivot_Analysis")
    ws_pivot.views.sheetView[0].showGridLines = True

    ws_pivot["B2"] = "CROSS-TABULATED EXPLORATORY PIVOT ANALYSIS"
    ws_pivot["B2"].font = title_font
    ws_pivot["B2"].fill = dark_fill
    ws_pivot.merge_cells("B2:L2")
    ws_pivot.row_dimensions[2].height = 28

    # Table 1: Top 15 Production Countries
    ws_pivot["B4"] = "Top 15 Production Countries (Primary Lead Country)"
    ws_pivot["B4"].font = section_font

    p1_headers = ["Rank", "Country", "Movies", "TV Shows", "Total Titles", "Movie Share %", "% of Catalog"]
    for idx, h in enumerate(p1_headers, start=2):
        col_let = get_column_letter(idx)
        ws_pivot[f"{col_let}5"] = h
        ws_pivot[f"{col_let}5"].font = header_font
        ws_pivot[f"{col_let}5"].fill = header_fill
        ws_pivot[f"{col_let}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws_pivot[f"{col_let}5"].border = thin_border
    ws_pivot.row_dimensions[5].height = 20

    top_countries = (
        df_clean[df_clean['primary_country'] != 'Unknown']['primary_country']
        .value_counts()
        .head(15)
        .index.tolist()
    )

    p_row = 6
    for rank, c_name in enumerate(top_countries, 1):
        ws_pivot[f"B{p_row}"] = rank
        ws_pivot[f"B{p_row}"].font = data_font
        ws_pivot[f"B{p_row}"].alignment = Alignment(horizontal="center")
        ws_pivot[f"B{p_row}"].border = thin_border

        ws_pivot[f"C{p_row}"] = c_name
        ws_pivot[f"C{p_row}"].font = data_font
        ws_pivot[f"C{p_row}"].border = thin_border

        ws_pivot[f"D{p_row}"] = f'=COUNTIFS(Cleaned_Data!$G$2:$G$8808, "{c_name}", Cleaned_Data!$B$2:$B$8808, "Movie")'
        ws_pivot[f"D{p_row}"].font = data_font
        ws_pivot[f"D{p_row}"].number_format = "#,##0"
        ws_pivot[f"D{p_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"D{p_row}"].border = thin_border

        ws_pivot[f"E{p_row}"] = f'=COUNTIFS(Cleaned_Data!$G$2:$G$8808, "{c_name}", Cleaned_Data!$B$2:$B$8808, "TV Show")'
        ws_pivot[f"E{p_row}"].font = data_font
        ws_pivot[f"E{p_row}"].number_format = "#,##0"
        ws_pivot[f"E{p_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"E{p_row}"].border = thin_border

        ws_pivot[f"F{p_row}"] = f"=D{p_row}+E{p_row}"
        ws_pivot[f"F{p_row}"].font = data_font
        ws_pivot[f"F{p_row}"].number_format = "#,##0"
        ws_pivot[f"F{p_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"F{p_row}"].border = thin_border

        ws_pivot[f"G{p_row}"] = f"=D{p_row}/F{p_row}"
        ws_pivot[f"G{p_row}"].font = data_font
        ws_pivot[f"G{p_row}"].number_format = "0.0%"
        ws_pivot[f"G{p_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"G{p_row}"].border = thin_border

        ws_pivot[f"H{p_row}"] = f"=F{p_row}/KPI_Summary!B6"
        ws_pivot[f"H{p_row}"].font = data_font
        ws_pivot[f"H{p_row}"].number_format = "0.0%"
        ws_pivot[f"H{p_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"H{p_row}"].border = thin_border

        if p_row % 2 == 1:
            for c in ["B", "C", "D", "E", "F", "G", "H"]:
                ws_pivot[f"{c}{p_row}"].fill = alt_row_fill
        p_row += 1

    # Table 2: Top 15 Primary Genres
    ws_pivot["J4"] = "Top 15 Primary Genres Distribution"
    ws_pivot["J4"].font = section_font

    g_headers = ["Rank", "Genre", "Movies", "TV Shows", "Total Titles"]
    for idx, h in enumerate(g_headers):
        col_let = chr(ord('J') + idx)
        ws_pivot[f"{col_let}5"] = h
        ws_pivot[f"{col_let}5"].font = header_font
        ws_pivot[f"{col_let}5"].fill = header_fill
        ws_pivot[f"{col_let}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws_pivot[f"{col_let}5"].border = thin_border

    top_genres = df_clean['primary_genre'].value_counts().head(15).index.tolist()
    g_row = 6
    for rank, g_name in enumerate(top_genres, 1):
        ws_pivot[f"J{g_row}"] = rank
        ws_pivot[f"J{g_row}"].font = data_font
        ws_pivot[f"J{g_row}"].alignment = Alignment(horizontal="center")
        ws_pivot[f"J{g_row}"].border = thin_border

        ws_pivot[f"K{g_row}"] = g_name
        ws_pivot[f"K{g_row}"].font = data_font
        ws_pivot[f"K{g_row}"].border = thin_border

        ws_pivot[f"L{g_row}"] = f'=COUNTIFS(Cleaned_Data!$R$2:$R$8808, "{g_name}", Cleaned_Data!$B$2:$B$8808, "Movie")'
        ws_pivot[f"L{g_row}"].font = data_font
        ws_pivot[f"L{g_row}"].number_format = "#,##0"
        ws_pivot[f"L{g_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"L{g_row}"].border = thin_border

        ws_pivot[f"M{g_row}"] = f'=COUNTIFS(Cleaned_Data!$R$2:$R$8808, "{g_name}", Cleaned_Data!$B$2:$B$8808, "TV Show")'
        ws_pivot[f"M{g_row}"].font = data_font
        ws_pivot[f"M{g_row}"].number_format = "#,##0"
        ws_pivot[f"M{g_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"M{g_row}"].border = thin_border

        ws_pivot[f"N{g_row}"] = f"=L{g_row}+M{g_row}"
        ws_pivot[f"N{g_row}"].font = data_font
        ws_pivot[f"N{g_row}"].number_format = "#,##0"
        ws_pivot[f"N{g_row}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"N{g_row}"].border = thin_border

        if g_row % 2 == 1:
            for c in ["J", "K", "L", "M", "N"]:
                ws_pivot[f"{c}{g_row}"].fill = alt_row_fill
        g_row += 1

    # Table 3: TV Show Longevity / Season Distribution
    start_s_row = p_row + 2
    ws_pivot[f"B{start_s_row}"] = "TV Show Longevity (Season Count Distribution)"
    ws_pivot[f"B{start_s_row}"].font = section_font

    s_headers = ["Seasons Count", "TV Shows Count", "% of All TV Shows", "Cumulative %"]
    for idx, h in enumerate(s_headers, start=2):
        col_let = get_column_letter(idx)
        ws_pivot[f"{col_let}{start_s_row+1}"] = h
        ws_pivot[f"{col_let}{start_s_row+1}"].font = header_font
        ws_pivot[f"{col_let}{start_s_row+1}"].fill = header_fill
        ws_pivot[f"{col_let}{start_s_row+1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_pivot[f"{col_let}{start_s_row+1}"].border = thin_border

    season_brackets = [
        ("1 Season", 1),
        ("2 Seasons", 2),
        ("3 Seasons", 3),
        ("4 Seasons", 4),
        ("5 Seasons", 5),
        ("6+ Seasons", ">=6")
    ]

    s_curr = start_s_row + 2
    for s_label, s_val in season_brackets:
        ws_pivot[f"B{s_curr}"] = s_label
        ws_pivot[f"B{s_curr}"].font = data_font
        ws_pivot[f"B{s_curr}"].border = thin_border

        if s_val == ">=6":
            ws_pivot[f"C{s_curr}"] = '=COUNTIF(Cleaned_Data!$Q$2:$Q$8808, ">=6")'
        else:
            ws_pivot[f"C{s_curr}"] = f'=COUNTIF(Cleaned_Data!$Q$2:$Q$8808, {s_val})'
        ws_pivot[f"C{s_curr}"].font = data_font
        ws_pivot[f"C{s_curr}"].number_format = "#,##0"
        ws_pivot[f"C{s_curr}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"C{s_curr}"].border = thin_border

        ws_pivot[f"D{s_curr}"] = f"=C{s_curr}/KPI_Summary!F6"
        ws_pivot[f"D{s_curr}"].font = data_font
        ws_pivot[f"D{s_curr}"].number_format = "0.0%"
        ws_pivot[f"D{s_curr}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"D{s_curr}"].border = thin_border

        if s_curr == start_s_row + 2:
            ws_pivot[f"E{s_curr}"] = f"=D{s_curr}"
        else:
            ws_pivot[f"E{s_curr}"] = f"=E{s_curr-1}+D{s_curr}"
        ws_pivot[f"E{s_curr}"].font = data_font
        ws_pivot[f"E{s_curr}"].number_format = "0.0%"
        ws_pivot[f"E{s_curr}"].alignment = Alignment(horizontal="right")
        ws_pivot[f"E{s_curr}"].border = thin_border

        if s_curr % 2 == 1:
            for c in ["B", "C", "D", "E"]:
                ws_pivot[f"{c}{s_curr}"].fill = alt_row_fill
        s_curr += 1

    pivot_widths = {"A": 4, "B": 10, "C": 20, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 6, "J": 8, "K": 26, "L": 14, "M": 14, "N": 14}
    for col, width in pivot_widths.items():
        ws_pivot.column_dimensions[col].width = width

    # ----------------------------------------------------
    # 3. SHEET: Data_Dictionary
    # ----------------------------------------------------
    print("Building Data_Dictionary sheet...")
    ws_dict = wb.create_sheet(title="Data_Dictionary")
    ws_dict.views.sheetView[0].showGridLines = True

    ws_dict["B2"] = "NETFLIX TITLES DATA DICTIONARY & METADATA"
    ws_dict["B2"].font = title_font
    ws_dict["B2"].fill = dark_fill
    ws_dict.merge_cells("B2:G2")
    ws_dict.row_dimensions[2].height = 28

    dict_headers = ["Field Name", "Data Type", "Role", "Null Count", "Description & Cleaning Notes", "Example Value"]
    for idx, h in enumerate(dict_headers, start=2):
        col_let = get_column_letter(idx)
        ws_dict[f"{col_let}4"] = h
        ws_dict[f"{col_let}4"].font = header_font
        ws_dict[f"{col_let}4"].fill = header_fill
        ws_dict[f"{col_let}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dict[f"{col_let}4"].border = thin_border
    ws_dict.row_dimensions[4].height = 22

    dict_entries = [
        ("show_id", "VARCHAR(10)", "Primary Key", "0", "Unique asset identifier for each movie or TV show", "s1"),
        ("type", "VARCHAR(10)", "Categorical Dimension", "0", "Type of content: strictly 'Movie' or 'TV Show'", "Movie"),
        ("title", "VARCHAR(255)", "Descriptive Attribute", "0", "Official release title, trimmed of extra whitespace", "Dick Johnson Is Dead"),
        ("director", "TEXT", "Creator Dimension", "0", "Director names; 2,634 raw missing values imputed with 'Unknown'", "Kirsten Johnson"),
        ("cast", "TEXT", "Talent Dimension", "0", "Lead and supporting actors; 825 missing values imputed with 'Unknown'", "Vicky Kaushal"),
        ("country", "VARCHAR(255)", "Geographic Dimension", "0", "Production countries; 831 missing values imputed with 'Unknown'", "United States, India"),
        ("primary_country", "VARCHAR(100)", "Lead Geographic Dimension", "0", "Engineered lead country before first comma for clean aggregations", "United States"),
        ("date_added", "DATE", "Temporal Dimension", "10", "Standardized ISO date (YYYY-MM-DD) when title debuted on Netflix", "2021-09-25"),
        ("year_added", "INTEGER", "Temporal Dimension", "10", "Calendar year title was added to Netflix (2008-2021)", "2021"),
        ("month_added", "INTEGER", "Temporal Dimension", "10", "Numerical month added (1 to 12) for seasonal analysis", "9"),
        ("month_name_added", "VARCHAR(20)", "Temporal Dimension", "10", "Full month name title was added", "September"),
        ("release_year", "INTEGER", "Temporal Dimension", "0", "Original theatrical or broadcast release year (1925-2021)", "2020"),
        ("rating", "VARCHAR(10)", "Certification Dimension", "0", "Age maturity certification. 3 misplaced records fixed; missing imputed to 'NR'", "PG-13"),
        ("rating_category", "VARCHAR(50)", "Audience Grouping", "0", "Consolidated maturity tier: Adults (18+), Teens, Older Kids, Kids, Unrated", "Teens (13-17)"),
        ("duration", "VARCHAR(50)", "Descriptive Attribute", "0", "Textual duration specification ('90 min' or '2 Seasons')", "90 min"),
        ("duration_minutes", "INTEGER", "Numerical Metric", "2,676 (TV Shows)", "Integer runtime in minutes for Movies; NULL for TV Shows", "90"),
        ("duration_seasons", "INTEGER", "Numerical Metric", "6,131 (Movies)", "Integer season count for TV Shows; NULL for Movies", "2"),
        ("listed_in", "VARCHAR(255)", "Categorical Dimension", "0", "Full comma-separated genre classifications from Netflix", "Documentaries"),
        ("primary_genre", "VARCHAR(100)", "Lead Category Dimension", "0", "Engineered first listed genre for macro-level category analysis", "Documentaries"),
        ("description", "TEXT", "Editorial Attribute", "0", "Official synopsis summary describing title premise", "As her father nears the end of...")
    ]

    d_row = 5
    for f_name, d_type, role, null_cnt, desc, ex_val in dict_entries:
        ws_dict[f"B{d_row}"] = f_name
        ws_dict[f"B{d_row}"].font = Font(name=font_family, size=9, bold=True, color="1F2937")
        ws_dict[f"B{d_row}"].border = thin_border

        ws_dict[f"C{d_row}"] = d_type
        ws_dict[f"C{d_row}"].font = data_font
        ws_dict[f"C{d_row}"].border = thin_border

        ws_dict[f"D{d_row}"] = role
        ws_dict[f"D{d_row}"].font = data_font
        ws_dict[f"D{d_row}"].border = thin_border

        ws_dict[f"E{d_row}"] = null_cnt
        ws_dict[f"E{d_row}"].font = data_font
        ws_dict[f"E{d_row}"].border = thin_border

        ws_dict[f"F{d_row}"] = desc
        ws_dict[f"F{d_row}"].font = data_font
        ws_dict[f"F{d_row}"].border = thin_border

        ws_dict[f"G{d_row}"] = ex_val
        ws_dict[f"G{d_row}"].font = data_font
        ws_dict[f"G{d_row}"].border = thin_border

        if d_row % 2 == 1:
            for c in ["B", "C", "D", "E", "F", "G"]:
                ws_dict[f"{c}{d_row}"].fill = alt_row_fill
        d_row += 1

    dict_widths = {"A": 4, "B": 20, "C": 18, "D": 22, "E": 18, "F": 45, "G": 24}
    for col, width in dict_widths.items():
        ws_dict.column_dimensions[col].width = width

    # ----------------------------------------------------
    # 4. SHEET: Cleaned_Data (Fast append)
    # ----------------------------------------------------
    print("Writing Cleaned_Data sheet (8,807 rows)...")
    ws_clean = wb.create_sheet(title="Cleaned_Data")
    ws_clean.views.sheetView[0].showGridLines = True

    # Use dataframe_to_rows for high-performance streaming
    clean_cols = df_clean.columns.tolist()
    for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=True), start=1):
        ws_clean.append(row)
        if r_idx == 1:
            for col_idx in range(1, len(clean_cols) + 1):
                cell = ws_clean.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_clean.row_dimensions[1].height = 22
    ws_clean.auto_filter.ref = f"A1:{get_column_letter(len(clean_cols))}{len(df_clean)+1}"
    for col_idx, col_name in enumerate(clean_cols, 1):
        col_let = get_column_letter(col_idx)
        ws_clean.column_dimensions[col_let].width = min(max(len(str(col_name)), 10) + 4, 32)

    # ----------------------------------------------------
    # 5. SHEET: Raw_Data (Fast append)
    # ----------------------------------------------------
    print("Writing Raw_Data sheet (8,807 rows)...")
    ws_raw = wb.create_sheet(title="Raw_Data")
    ws_raw.views.sheetView[0].showGridLines = True

    raw_cols = df_raw.columns.tolist()
    for r_idx, row in enumerate(dataframe_to_rows(df_raw, index=False, header=True), start=1):
        ws_raw.append(row)
        if r_idx == 1:
            for col_idx in range(1, len(raw_cols) + 1):
                cell = ws_raw.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = sub_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_raw.row_dimensions[1].height = 22
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(raw_cols))}{len(df_raw)+1}"
    for col_idx, col_name in enumerate(raw_cols, 1):
        col_let = get_column_letter(col_idx)
        ws_raw.column_dimensions[col_let].width = min(max(len(str(col_name)), 10) + 4, 30)

    # Save
    print(f"Saving finalized workbook to {excel_path}...")
    wb.save(excel_path)
    print("Excel workbook build complete and saved successfully!")

if __name__ == '__main__':
    build_netflix_excel()
